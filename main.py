import tkinter as tk
from tkinter import ttk
import openpyxl

def toggle_mode():
    if mode_switch.instate(["selected"]):
        style.theme_use("forest-light")
    else:
        style.theme_use("forest-dark")

def insert_row():
    name = name_entry.get()
    id = int(Id_entry.get())
    role = role_combo.get()
    participation = "yes" if a.get() else "no"

    print(name, id, role, participation)

    path = "Book1.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    row_values = [name, id, role, participation]
    sheet.append(row_values)
    workbook.save(path)

    treeview.insert('', tk.END, values=row_values)

    name_entry.delete(0, "end")
    name_entry.insert(0, "Name")
    Id_entry.delete(0, "end")
    Id_entry.insert(0, "Age")
    role_combo.set(listt[0])
    check.state(["!selected"])

def delete_row():
    
    col = del_choice.get()
    ind = del_choice.current()
    val = delval_entry.get()

    print(f"deleting:({col}, {ind},{val})")

    path = "Book1.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    
    row_to_delete = None
    count=2
    for row in sheet.iter_rows(min_row=2, max_col=4, max_row=sheet.max_row, values_only=True):
        
        if str(row[ind]) == val:
            row_to_delete=count
            break
        else:
            count+=1
            pass
    print(row_to_delete)
    if row_to_delete is not None:
        sheet.delete_rows(row_to_delete)
        print(f"Row with {val} deleted successfully.")
    else:
        print(f"Row with {val} not found.")
    
    workbook.save(path)

    
    treeview.delete(treeview.get_children()[count-2])
    
    delval_entry.delete(0, "end")
    delval_entry.insert(0, "Enter Value")
    del_choice.set(listt[0])

def update_row():
    old_col_ind = up_choice1.current()
    old_val = upval_entry1.get()
    new_col_ind = new_choice.current()
    new_val = new_entry.get()
    
    path = "Book1.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    flag=0
    count=2
    new_values=[]
    for row in sheet.iter_rows(min_row=2, max_col=4, max_row=sheet.max_row, values_only=True):
        
        if str(row[old_col_ind]) == old_val:
            row = list(row)
            row[new_col_ind]=new_val
            row=tuple(row)
            flag=1
            for i in range(0,4):
                new_values.append(row[i])
            break
        else:
            count+=1
            pass
    sheet.insert_rows(count)
    for i, value in enumerate(new_values, start=1):
        sheet.cell(row=count, column=i, value=value)
    sheet.delete_rows(count+1)
    if flag==1:
        print("Value updated")
    else:
        printf("Value not found")
    workbook.save(path)



    treeview.delete(treeview.get_children()[count-2])
    treeview.insert("", count-2, values=new_values)
    

    upval_entry1.delete(0, "end")
    upval_entry1.insert(0, "Enter Value")
    new_entry.delete(0,"end")
    new_entry.insert(0,"Enter new Value")
    up_choice1.set(listt1[0])
    new_choice.set(listt1[0])

 


def load_data():
    path = "Book1.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    list_values = list(sheet.values)
    print(list_values)  
    for col_name in list_values[0]:
        treeview.heading(col_name, text=col_name)

    for value_tuple in list_values[1:]:
        treeview.insert('', tk.END, values=value_tuple)


root = tk.Tk() #this is our root widget or window, so even if we just write this apart from any other code, we just get a plain blank window
root.title("Excel-lent")
style = ttk.Style(root)
root.tk.call("source", "forest-light.tcl")
root.tk.call("source", "forest-dark.tcl")
style.theme_use("forest-dark")

#defining the frame in the root window so we can enter the widgets in a designated location
frame = ttk.Frame(root)
frame.pack()

#first region where we can do operations
insert_frame = ttk.LabelFrame(frame,text="Insert Row")
insert_frame.grid(row=0,column=0,padx=20,pady=10)

name_entry = ttk.Entry(insert_frame)
name_entry.insert(0,"Name")
name_entry.bind("<FocusIn>",lambda e:name_entry.delete('0','end'))
name_entry.grid(row=0,column=0,sticky="ew",padx=20,pady=10)

Id_entry = ttk.Spinbox(insert_frame, from_=1,to=1000)
Id_entry.insert(0,"Id")
Id_entry.bind("<FocusIn>",lambda e:Id_entry.delete('0','end'))
Id_entry.grid(row=1,column=0,sticky="ew",padx=20,pady=10)

listt = ['Core Member','Cocom Member','OC Member','Other']
role_combo = ttk.Combobox(insert_frame,values=listt)
role_combo.current(0)
role_combo.grid(row=2,column=0,sticky="ew",padx=20,pady=10)

a=tk.BooleanVar()
check = ttk.Checkbutton(insert_frame,text = "Participating",variable=a)
check.grid(row=3,column=0,sticky = "nsew",padx=20,pady=10)

button = tk.Button(insert_frame,text = "Insert",command=insert_row)
button.grid(row=4,column=0,sticky = "nsew",padx=20,pady=10)

delete_frame = ttk.LabelFrame(frame,text="Delete Row")
delete_frame.grid(row=1,column=0,padx=20,pady=10)

del_label = ttk.Label(delete_frame,text="Enter the column to use for reference")
del_label.grid(row=0,column=0,sticky="ew",padx=20,pady=10)

listt1 = ['Name','Id','Role','Participation']
del_choice = ttk.Combobox(delete_frame,values=listt1)
del_choice.current(0)
del_choice.grid(row=1,column=0,sticky="ew",padx=20,pady=10)

delval_entry = ttk.Entry(delete_frame)
delval_entry.insert(0,"Enter Value")
delval_entry.bind("<FocusIn>",lambda e:delval_entry.delete('0','end'))
delval_entry.grid(row=2,column=0,sticky="ew",padx=20,pady=10)

button = tk.Button(delete_frame,text = "Delete",command=delete_row)
button.grid(row=3,column=0,sticky = "nsew",padx=20,pady=10)

separator = ttk.Separator(delete_frame)
separator.grid(row=4, column=0, padx=(20, 10), pady=10, sticky="ew")

mode_switch = ttk.Checkbutton(
    delete_frame, text="Mode", style="Switch", command=toggle_mode)
mode_switch.grid(row=5, column=0, padx=5, pady=10, sticky="nsew")

Update_frame = ttk.LabelFrame(frame,text="Update Row")
Update_frame.grid(row=1,column=1,padx=20,pady=10)

up_label = ttk.Label(Update_frame,text="Enter a column to use for reference")
up_label.grid(row=0,column=0,sticky="ew",padx=20,pady=10)

listt2 = ['Name','Id','Role','Participation']
up_choice1 = ttk.Combobox(Update_frame,values=listt1)
up_choice1.current(0)
up_choice1.grid(row=1,column=0,sticky="ew",padx=20,pady=10)

upval_entry1 = ttk.Entry(Update_frame)
upval_entry1.insert(0,"Enter Value")
upval_entry1.bind("<FocusIn>",lambda e:upval_entry1.delete('0','end'))
upval_entry1.grid(row=2,column=0,sticky="ew",padx=20,pady=10)

new_label = ttk.Label(Update_frame,text="Enter the column of the new value")
new_label.grid(row=0,column=1,sticky="ew",padx=20,pady=10)

new_choice = ttk.Combobox(Update_frame,values=listt1)
new_choice.current(0)
new_choice.grid(row=1,column=1,sticky="ew",padx=20,pady=10)

new_entry = ttk.Entry(Update_frame)
new_entry.insert(0,"Enter Value")
new_entry.bind("<FocusIn>",lambda e:new_entry.delete('0','end'))
new_entry.grid(row=2,column=1,sticky="ew",padx=20,pady=10)

button = tk.Button(Update_frame,text = "Update",command=update_row)
button.grid(row=3,column=0,columnspan=2,sticky = "nsew",padx=20,pady=10)


treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0, column=1, pady=10)
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="y")

cols = ("Name", "ID", "Role", "Participation")
treeview = ttk.Treeview(treeFrame, show="headings",
                        yscrollcommand=treeScroll.set, columns=cols, height=13)
treeview.column("Name", width=100)
treeview.column("ID", width=100)
treeview.column("Role", width=120)
treeview.column("Participation", width=100)
treeview.pack()
treeScroll.config(command=treeview.yview)
load_data()
root.mainloop()# event loop that launches our app and keeps executing till we close our app by terminating it